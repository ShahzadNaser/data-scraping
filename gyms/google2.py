from selenium import webdriver
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import time
driver = webdriver.Firefox()
for sht in ["riverside","oakland","long_beach","bakersfield"]:
    wb = load_workbook(filename = '/home/shahzad/Documents/personal_work/gyms/{0}.xlsx'.format(sht))
    sheet = wb.active
    i=0
    while i < sheet.max_row:
        i+=1
        key1 = "A{0}".format(i)
        key2 = "B{0}".format(i)
        url = "https://www.google.com/search?newwindow=1&client=ubuntu&channel=fs&sxsrf=ACYBGNSYrUj2jMGc8RF8Pftl4AxN-HXV-w%3A1569642042937&ei=OtaOXcHlOIvWwAK80Y3YCw&q={0} {1}&gs_l=psy-ab.3..35i39.318958.319876..320892...0.2..0.369.639.2-1j1......0....1..gws-wiz.......0i71j35i304i39.Gbvdc1NDLk4&ved=0ahUKEwjB6pKjzPLkAhULK1AKHbxoA7sQ4dUDCAo&uact=5".format(sheet[key1].value,sheet[key2].value)
        try:
            driver.get(url)
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            search = soup.find('div', {'id':'search'}).find("div", {'class':'srg'})
            links = search.find_all("a", href=True)
            yelp_link = ""
            for link in links:
                if("yelp.com" in link["href"]):
                    yelp_link = link["href"]
                    break 
            key8 = "H{0}".format(i)
            sheet[key8].value = yelp_link
            wb.save(filename = '/home/shahzad/Documents/personal_work/gyms/{0}.xlsx'.format(sht))
            time.sleep(5)
        except Exception:
            pass

