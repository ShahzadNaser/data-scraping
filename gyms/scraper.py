from selenium import webdriver
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import time
driver = webdriver.Firefox()
for sht in ["riverside","oakland","long_beach","bakersfield"]:
    wb = load_workbook(filename = '/home/shahzad/Documents/personal_work/gyms/{0}.xlsx'.format(sht))
    sheet = wb.active
    i=0
    while i < sheet.max_row:
        i+=1
        key8 = "H{0}".format(i)
        url = sheet[key8].value
        try:
            driver.get(url)
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            name = phone = addr = site = owner =  ''
            name_h1 = soup.find('div', class_='biz-page-header-left claim-status')
            if(name_h1):
                if(name_h1.find_all("h1")):
                    for h1 in name_h1.find_all("h1"):
                        name = name+" "+h1.text
            text_box = soup.find('div', class_='mapbox-text')
            if(text_box):
                addr_span = text_box.find("strong",{"class":"street-address"})
                if(addr_span):
                    addr = addr_span.text.replace("\n","").replace("        ","")
                phone_span = text_box.find("span",{"class":"biz-phone"})
                if(phone_span):
                    phone = phone_span.text.replace("\n","").replace("            ","")
                site_span = soup.find('div', class_='mapbox-text').find("span",{"class":"biz-website"})
                if(site_span):
                    site = "http://"+site_span.find("a").text
            buisness_div = soup.find('div', class_='meet-business-owner')
            if(buisness_div):
                owner = buisness_div.find("span",{"class":"user-display-name"})
                if(owner):
                    owner = owner.text.replace("\n","").replace("                        ","")
            key1 = "C{0}".format(i)
            key2 = "D{0}".format(i)
            key3 = "E{0}".format(i)
            key4 = "F{0}".format(i)
            key5 = "G{0}".format(i)
            sheet[key1].value = name
            sheet[key2].value = addr
            sheet[key3].value = phone
            sheet[key4].value = site
            sheet[key5].value = owner
            wb.save(filename = '/home/shahzad/Documents/personal_work/gyms/{0}.xlsx'.format(sht))
            time.sleep(5)
        except Exception:
            pass

