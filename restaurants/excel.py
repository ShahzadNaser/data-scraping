from openpyxl import load_workbook
for sht in ["attock","dgh","dih","gujrat","haripur","kasur","jhang","mansehra","mianwali","taxila"]:
# for sht in ["gujrat1"]:
    wb = load_workbook(filename = '/home/shahzad/Documents/personal_work/restaurants/{0}.xlsx'.format(sht))
    sheet = wb.active
    i=0
    while i < sheet.max_row:
        time_arr = []
        i+=1
        key8 = "H{0}".format(i)
        if(sheet[key8].value):
            time_arr = sheet[key8].value.splitlines()
            if(time_arr and len(time_arr)>0):
                sheet[key8].value = time_arr[0].replace("Closed","").replace("Friday","").replace("Saturday","").replace("Sunday","").replace("Monday","").replace("Tuesday","").replace("Wednesday","").replace("Thursday","")
    wb.save(filename = '/home/shahzad/Documents/personal_work/restaurants/{0}.xlsx'.format(sht))
