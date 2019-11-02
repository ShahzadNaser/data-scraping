from selenium import webdriver
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import time
driver = webdriver.Firefox()
for sht in ["wahcantt"]:
    wb = load_workbook(filename = '/home/shahzad/Documents/personal_work/restaurants/{0}.xlsx'.format(sht))
    sheet = wb.active
    i=0
    while i < sheet.max_row:
        i+=1
        key1 = "A{0}".format(i)
        key2 = "B{0}".format(i)
        url = "https://www.google.com/search?client=ubuntu&channel=fs&ei=mNJ7Xd_6LsfNwQLK4I7gCg&q={0} {1}".format(sheet[key1].value,sheet[key2].value)
        try:
            driver.get(url)
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            fblink = phone = addr = timing = cat = pr =  ''
            if(len(soup.find_all('div', class_='i4J0ge'))>0):
                if(soup.find('div', class_='i4J0ge').find("div",{"data-attrid":"kc:/location/location:address"}).find("span",{"class":"LrzXr"})):
                    addr = soup.find('div', class_='i4J0ge').find("div",{"data-attrid":"kc:/location/location:address"}).find("span",{"class":"LrzXr"}).text
                if(soup.find('div', class_='fYOrjf kp-hc').find("div",{"data-attrid":"kc:/local:one line summary"})):
                    spans = soup.find('div', class_='fYOrjf kp-hc').find("div",{"data-attrid":"kc:/local:one line summary"}).find_all("span")
                    if(len(spans)>1):
                        pr = spans[0].text
                        cat = spans[1].text
                    elif(len(spans)==1):
                        cat = spans[0].text
                if(soup.find('div', class_='i4J0ge').find("span",{"data-local-attribute":"d3ph"})):
                    phone = soup.find('div', class_='i4J0ge').find("span",{"data-local-attribute":"d3ph"}).find('span').text
                if(soup.find('div', class_='i4J0ge').find("div",{"data-attrid":"kc:/location/location:hours"})):
                    tbldiv = soup.find('div', class_='i4J0ge').find("div",{"data-attrid":"kc:/location/location:hours"})
                    if(tbldiv.find("table").find("tbody").find_all("tr")):
                        for tr in tbldiv.find("table").find("tbody").find_all("tr"):
                            td = tr.find_all("td")
                            if(len(td)>1):
                                timing += td[0].text+" "+td[1].text+"\n"
                            elif(len(td)==1):
                                timing += td[0].text+"\n"
            search = soup.find('div', {'id':'search'}).find("div", {'class':'srg'})
            links = search.find_all("a", href=True)
            for link in links:
                if("facebook.com" in link["href"]):
                    fblink = link["href"]
                    break    
            key3 = "C{0}".format(i)
            key4 = "D{0}".format(i)
            key5 = "E{0}".format(i)
            key6 = "F{0}".format(i)
            key7 = "G{0}".format(i)
            key8 = "H{0}".format(i)
            sheet[key4].value = phone
            sheet[key3].value = addr
            sheet[key5].value = cat
            sheet[key6].value = pr
            sheet[key7].value = fblink
            sheet[key8].value = timing
            wb.save(filename = '/home/shahzad/Documents/personal_work/restaurants/{0}.xlsx'.format(sht))
            time.sleep(7)
        except Exception:
            pass

